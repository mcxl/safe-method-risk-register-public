from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


VISIBLE_SHEETS = [
    "Cover",
    "Dashboard",
    "HRCW Register",
    "Hold Points Schedule",
    "Risk Register",
]
EXPECTED_SHEETS = [*VISIBLE_SHEETS, "Lists"]
ERROR_RE = re.compile(r"#(REF!|DIV/0!|VALUE!|NAME\\?|N/A|NUM!|NULL!)")


def assert_phase5b_xlsx(path: Path, document_set: dict[str, Any]) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False)
    failures: list[str] = []

    if wb.sheetnames != EXPECTED_SHEETS:
        failures.append(f"Workbook sheets {wb.sheetnames} != {EXPECTED_SHEETS}")

    for sheet_name in VISIBLE_SHEETS:
        ws = wb[sheet_name]
        if ws.sheet_state != "visible":
            failures.append(f"{sheet_name} is not visible")
        if ws.page_setup.paperSize != 9:
            failures.append(f"{sheet_name} is not A4")
        if ws.page_setup.orientation != "landscape":
            failures.append(f"{sheet_name} is not landscape")
        if int(ws.page_setup.fitToWidth or 0) != 1:
            failures.append(f"{sheet_name} fitToWidth is not 1")
        if int(ws.page_setup.fitToHeight or 0) != 0:
            failures.append(f"{sheet_name} fitToHeight is not 0")
        if "Page &P of &N" not in (ws.oddFooter.right.text or ""):
            failures.append(f"{sheet_name} footer lacks page numbers")

    if wb["Lists"].sheet_state != "hidden":
        failures.append("Lists sheet is not hidden")

    if "RiskRatings" not in wb.defined_names:
        failures.append("RiskRatings defined name is missing")
    if "HrcwTriggerValues" not in wb.defined_names:
        failures.append("HrcwTriggerValues defined name is missing")

    hrcw = wb["HRCW Register"]
    risk = wb["Risk Register"]
    dashboard = wb["Dashboard"]

    if hrcw["C4"].value not in {"YES", "COND", "NO"}:
        failures.append("HRCW Triggered column does not start with a dropdown value")
    if not any("HrcwTriggerValues" in str(dv.formula1) for dv in hrcw.data_validations.dataValidation):
        failures.append("HRCW trigger data validation is missing")

    expected_risk_rows = len(document_set["risk_register"])
    expected_hrcw_rows = len(document_set["hrcw_register"])
    expected_hp_rows = len(document_set["hold_point_schedule"])
    if risk.max_row < expected_risk_rows + 3:
        failures.append("Risk Register does not contain expected risk rows")
    if hrcw.max_row < expected_hrcw_rows + 3:
        failures.append("HRCW Register does not contain expected HRCW rows")
    if wb["Hold Points Schedule"].max_row < expected_hp_rows + 3:
        failures.append("Hold Points Schedule does not contain expected hold-point rows")

    for cell in ("K4", "L4", "M4"):
        if not is_formula(risk[cell].value):
            failures.append(f"Risk Register {cell} is not a formula")
    for cell in ("B4", "B5", "B6", "B16", "B17", "C4", "C17"):
        if not is_formula(dashboard[cell].value):
            failures.append(f"Dashboard {cell} is not a formula")

    formula_cells = [
        cell
        for sheet in wb.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if is_formula(cell.value)
    ]
    minimum_formula_count = expected_risk_rows * 3 + 14
    if len(formula_cells) < minimum_formula_count:
        failures.append(
            f"Formula count {len(formula_cells)} is below expected minimum {minimum_formula_count}"
        )

    if not any("RiskRatings" in str(dv.formula1) for dv in risk.data_validations.dataValidation):
        failures.append("Risk rating data validation is missing")

    if hrcw["A3"].fill.fgColor.rgb[-6:] != "1F3864":
        failures.append("HRCW header is not navy")
    if risk["F4"].fill.fgColor.rgb[-6:] not in {"FFCCCC", "FFE5B4", "C6EFCE"}:
        failures.append("Risk rating cell does not have traffic-light fill")
    if "DRAFT" not in str(wb["Cover"]["A1"].value):
        failures.append("Cover does not mark workbook as DRAFT")

    errors = scan_formula_error_literals(wb)
    if errors:
        failures.append(f"Workbook contains formula error literals: {errors[:5]}")

    return {
        "status": "pass" if not failures else "fail",
        "workbook": str(path),
        "formula_count": len(formula_cells),
        "sheet_names": wb.sheetnames,
        "failures": failures,
    }


def scan_recalculated_values(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    errors = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and ERROR_RE.search(value):
                    errors.append(f"{sheet.title}!{cell.coordinate}={value}")
    return {"status": "pass" if not errors else "fail", "errors": errors}


def scan_formula_error_literals(wb) -> list[str]:
    errors = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and ERROR_RE.search(value):
                    errors.append(f"{sheet.title}!{cell.coordinate}={value}")
    return errors


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Assert Safe Method Phase 5B XLSX output.")
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--document-set-json")
    parser.add_argument("--scan-values", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.scan_values:
        report = scan_recalculated_values(Path(args.workbook))
    else:
        if not args.document_set_json:
            raise SystemExit("--document-set-json is required unless --scan-values is used")
        document_set = json.loads(Path(args.document_set_json).read_text(encoding="utf-8"))
        report = assert_phase5b_xlsx(Path(args.workbook), document_set)

    print(json.dumps(report, indent=2))
    if report["status"] != "pass":
        raise SystemExit(1)


if __name__ == "__main__":
    main()
