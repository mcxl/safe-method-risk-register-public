from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


XLSX_RENDERER_VERSION = "phase5b.xlsx-renderer.v1"
MAX_EDIT_ROWS = 200
VISIBLE_SHEETS = [
    "Cover",
    "Dashboard",
    "HRCW Register",
    "Hold Points Schedule",
    "Risk Register",
]
ALL_SHEETS = [*VISIBLE_SHEETS, "Lists"]

COLORS = {
    "navy": "1F3864",
    "white": "FFFFFF",
    "border": "BFBFBF",
    "subtext": "404040",
    "note": "F2F2F2",
    "high": "FFCCCC",
    "medium": "FFE5B4",
    "low": "C6EFCE",
    "yes_fill": "C6EFCE",
    "yes_text": "375623",
    "cond_fill": "FFF2CC",
    "cond_text": "7F6000",
    "no_fill": "D9D9D9",
    "no_text": "808080",
}


def render_xlsx(document_set: dict[str, Any], output_path: Path, filename: str) -> dict[str, Any]:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name in ALL_SHEETS:
        wb.create_sheet(sheet_name)

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    create_lists_sheet(wb)
    create_cover_sheet(wb["Cover"], document_set)
    create_hrcw_sheet(wb["HRCW Register"], document_set)
    create_hold_points_sheet(wb["Hold Points Schedule"], document_set)
    create_risk_register_sheet(wb["Risk Register"], document_set)
    create_dashboard_sheet(wb["Dashboard"], document_set)

    for sheet_name in VISIBLE_SHEETS:
        apply_print_setup(wb[sheet_name], filename)
    wb["Lists"].sheet_state = "hidden"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "renderer_version": XLSX_RENDERER_VERSION,
        "workflow_state": "DRAFT",
        "issue_ready": False,
        "output_path": str(output_path),
        "filename": filename,
        "output_hash_sha256": sha256_file(output_path),
    }


def create_lists_sheet(wb: Workbook) -> None:
    ws = wb["Lists"]
    ws.sheet_view.showGridLines = False
    ws.append(["RiskRatings", "HrcwTriggerValues"])
    rows = [
        ["High (3)", "YES"],
        ["Medium (2)", "COND"],
        ["Low (1)", "NO"],
    ]
    for row in rows:
        ws.append(row)

    style_header_row(ws, 1, 2)
    for column in ("A", "B"):
        ws.column_dimensions[column].width = 22

    wb.defined_names.add(DefinedName("RiskRatings", attr_text="'Lists'!$A$2:$A$4"))
    wb.defined_names.add(DefinedName("HrcwTriggerValues", attr_text="'Lists'!$B$2:$B$4"))


def create_cover_sheet(ws, document_set: dict[str, Any]) -> None:
    project = document_set["project"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.merge_cells("A1:E1")
    ws["A1"] = "Master Project Register - DRAFT"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=COLORS["white"])
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["navy"])
    ws["A1"].alignment = Alignment(horizontal="center")

    rows = [
        ("Workflow State", "DRAFT - not issue-ready until consultant review/sign-off."),
        ("Project", project.get("project_name")),
        ("Site Address", project.get("site_address")),
        ("Principal Contractor", project.get("principal_contractor")),
        ("WHS Consultant", project.get("whs_consultant")),
        ("Jurisdiction", project.get("jurisdiction")),
        ("Document Reference", project.get("document_ref")),
        ("Revision", project.get("revision")),
        ("Issue Date", project.get("issue_date")),
        ("Review Date", project.get("review_date")),
        ("Document Level", document_set.get("document_level")),
        ("Trade Packages", join_list(project.get("trade_packages", []))),
        ("SWMS Benchmark Note", document_set.get("swms_benchmark_note")),
    ]
    ws.append([])
    ws.append(["Field", "Value"])
    style_header_row(ws, 3, 2)
    for field, value in rows:
        ws.append([field, safe_text(value)])

    ws["B4"].comment = Comment(
        "Draft outputs remain blocked from issue-ready status until deterministic gates and consultant sign-off pass.",
        "Safe Method",
    )
    style_used_range(ws)
    set_widths(ws, [28, 120, 14, 14, 14])


def create_dashboard_sheet(ws, document_set: dict[str, Any]) -> None:
    risk_count = len(document_set["risk_register"])
    hrcw_count = len(document_set["hrcw_register"])
    hold_point_count = len(document_set["hold_point_schedule"])
    first_data_row = 4
    risk_last = first_data_row + max(risk_count, 1) - 1
    hrcw_last = first_data_row + max(hrcw_count, 1) - 1
    hold_last = first_data_row + max(hold_point_count, 1) - 1

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.merge_cells("A1:F1")
    ws["A1"] = "Dashboard - Live Rollup"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=COLORS["white"])
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["navy"])
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])
    ws.append(["Metric", "Live Formula", "Current Value", "Source"])
    style_header_row(ws, 3, 4)

    rows = [
        (
            "Risk-line count",
            f'=COUNTA(\'Risk Register\'!$A${first_data_row}:$A${risk_last})',
            "Risk Register",
        ),
        (
            "Hold-point count",
            f'=COUNTA(\'Hold Points Schedule\'!$A${first_data_row}:$A${hold_last})',
            "Hold Points Schedule",
        ),
        (
            "HRCW YES count",
            f'=COUNTIF(\'HRCW Register\'!$C${first_data_row}:$C${hrcw_last},"YES")',
            "HRCW Register",
        ),
        (
            "HRCW COND count",
            f'=COUNTIF(\'HRCW Register\'!$C${first_data_row}:$C${hrcw_last},"COND")',
            "HRCW Register",
        ),
        (
            "HRCW NO count",
            f'=COUNTIF(\'HRCW Register\'!$C${first_data_row}:$C${hrcw_last},"NO")',
            "HRCW Register",
        ),
        (
            "Packages needing SWMS",
            f'=SUMPRODUCT((\'Risk Register\'!$B${first_data_row}:$B${risk_last}<>"")/COUNTIF(\'Risk Register\'!$B${first_data_row}:$B${risk_last},\'Risk Register\'!$B${first_data_row}:$B${risk_last}&""))',
            "Risk Register",
        ),
        (
            "Initial High",
            f'=COUNTIF(\'Risk Register\'!$F${first_data_row}:$F${risk_last},"High (3)")',
            "Risk Register",
        ),
        (
            "Initial Medium",
            f'=COUNTIF(\'Risk Register\'!$F${first_data_row}:$F${risk_last},"Medium (2)")',
            "Risk Register",
        ),
        (
            "Initial Low",
            f'=COUNTIF(\'Risk Register\'!$F${first_data_row}:$F${risk_last},"Low (1)")',
            "Risk Register",
        ),
        (
            "Residual High",
            f'=COUNTIF(\'Risk Register\'!$G${first_data_row}:$G${risk_last},"High (3)")',
            "Risk Register",
        ),
        (
            "Residual Medium",
            f'=COUNTIF(\'Risk Register\'!$G${first_data_row}:$G${risk_last},"Medium (2)")',
            "Risk Register",
        ),
        (
            "Residual Low",
            f'=COUNTIF(\'Risk Register\'!$G${first_data_row}:$G${risk_last},"Low (1)")',
            "Risk Register",
        ),
        (
            "Average Risk Reduction",
            f'=AVERAGE(\'Risk Register\'!$M${first_data_row}:$M${risk_last})',
            "Risk Register",
        ),
        (
            "% Lines Reduced",
            f'=COUNTIF(\'Risk Register\'!$M${first_data_row}:$M${risk_last},">0")/COUNT(\'Risk Register\'!$M${first_data_row}:$M${risk_last})',
            "Risk Register",
        ),
    ]

    for label, formula, source in rows:
        ws.append([label, formula, formula, source])

    for row in range(4, 4 + len(rows)):
        ws.cell(row, 2).font = Font(name="Calibri", size=9, color=COLORS["subtext"])
        ws.cell(row, 3).number_format = "0.00" if row == 16 else "0"
        if row == 17:
            ws.cell(row, 3).number_format = "0.0%"

    style_used_range(ws)
    set_widths(ws, [30, 72, 18, 26, 14, 14])


def create_hrcw_sheet(ws, document_set: dict[str, Any]) -> None:
    headers = ["Ref", "Schedule 1 Item", "Triggered", "Category", "Packages", "SWMS Required", "Basis / Notes"]
    ws.append(["HRCW Register - DRAFT"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell(ws["A1"])
    ws.append([])
    ws.append(headers)
    style_header_row(ws, 3, len(headers))

    for row in document_set["hrcw_register"]:
        trigger = display_trigger(row)
        ws.append(
            [
                row.get("ref"),
                row.get("schedule_1_item"),
                trigger,
                row.get("category_title"),
                join_list(row.get("packages", [])),
                "YES" if row.get("swms_required") == "YES" or trigger in {"YES", "COND"} else "NO",
                join_list(
                    [
                        join_list(row.get("basis_refs", [])),
                        row.get("condition"),
                        row.get("notes"),
                        row.get("risk_description"),
                    ]
                ),
            ]
        )
        apply_trigger_style(ws.cell(ws.max_row, 3), trigger)

    add_list_validation(ws, f"C4:C{max(MAX_EDIT_ROWS, ws.max_row)}", "HrcwTriggerValues")
    add_trigger_conditional_formats(ws, f"C4:C{max(MAX_EDIT_ROWS, ws.max_row)}")
    style_used_range(ws)
    set_widths(ws, [10, 16, 13, 48, 55, 16, 80])
    ws.freeze_panes = "A4"


def create_hold_points_sheet(ws, document_set: dict[str, Any]) -> None:
    headers = ["Ref", "Status", "Title", "Packages", "Release Criteria", "Authority", "Evidence"]
    ws.append(["Hold Points Schedule - DRAFT"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell(ws["A1"])
    ws.append([])
    ws.append(headers)
    style_header_row(ws, 3, len(headers))

    for row in document_set["hold_point_schedule"]:
        ws.append(
            [
                row.get("ref"),
                row.get("status"),
                row.get("title"),
                join_list(row.get("packages", [])),
                row.get("release_criteria") or row.get("precondition"),
                row.get("release_authority") or row.get("authority_text"),
                row.get("evidence_required"),
            ]
        )

    style_used_range(ws)
    set_widths(ws, [11, 24, 42, 40, 70, 38, 55])
    ws.freeze_panes = "A4"


def create_risk_register_sheet(ws, document_set: dict[str, Any]) -> None:
    headers = [
        "Ref",
        "Package",
        "Activity",
        "Hazard",
        "Controls",
        "Initial Risk",
        "Residual Risk",
        "Hold Points",
        "Responsible Do",
        "Responsible Verify",
        "Initial Score",
        "Residual Score",
        "Risk Reduction",
    ]
    ws.append(["Risk Register - DRAFT"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell(ws["A1"])
    ws.append([])
    ws.append(headers)
    style_header_row(ws, 3, len(headers))

    for row in document_set["risk_register"]:
        excel_row = ws.max_row + 1
        ws.append(
            [
                row.get("ref"),
                row.get("trade_package"),
                join_list([row.get("phase"), row.get("activity")]),
                row.get("hazard"),
                format_controls(row.get("controls", [])),
                display_risk(row.get("initial_risk")),
                display_risk(row.get("residual_risk")),
                join_list(row.get("hold_points", [])),
                row.get("responsible_do"),
                row.get("responsible_verify"),
                score_formula(f"F{excel_row}"),
                score_formula(f"G{excel_row}"),
                f'=IF(OR(K{excel_row}="",L{excel_row}=""),"",K{excel_row}-L{excel_row})',
            ]
        )
        apply_risk_style(ws.cell(excel_row, 6), row.get("initial_risk"))
        apply_risk_style(ws.cell(excel_row, 7), row.get("residual_risk"))

    last_row = max(MAX_EDIT_ROWS, ws.max_row)
    add_list_validation(ws, f"F4:G{last_row}", "RiskRatings")
    add_risk_conditional_formats(ws, f"F4:G{last_row}")
    for row_number in range(4, ws.max_row + 1):
        ws.cell(row_number, 11).number_format = "0"
        ws.cell(row_number, 12).number_format = "0"
        ws.cell(row_number, 13).number_format = "0"

    style_used_range(ws)
    set_widths(ws, [11, 30, 42, 44, 75, 14, 14, 18, 30, 30, 14, 14, 15])
    ws.freeze_panes = "A4"


def display_trigger(row: dict[str, Any]) -> str:
    status = row.get("trigger_status")
    if status == "confirmed_hrcw":
        return "YES"
    if status == "conditional_hrcw":
        return "COND"
    if status == "not_triggered":
        return "NO"
    triggered = row.get("triggered")
    if triggered == "YES":
        return "YES"
    if triggered == "CONDITIONAL":
        return "COND"
    return "NO"


def display_risk(value: Any) -> str:
    text = safe_text(value)
    scores = {"High": "3", "Medium": "2", "Low": "1"}
    if text in scores:
        return f"{text} ({scores[text]})"
    return text


def score_formula(cell_ref: str) -> str:
    return f'=IF({cell_ref}="","",VALUE(MID({cell_ref},FIND("(",{cell_ref})+1,FIND(")",{cell_ref})-FIND("(",{cell_ref})-1)))'


def format_controls(controls: list[dict[str, Any]]) -> str:
    return "\n".join(
        join_list(
            [
                control.get("text"),
                f"Sources: {join_list(control.get('source_ids', []))}" if control.get("source_ids") else None,
                f"Status: {control.get('control_status')}" if control.get("control_status") else None,
            ]
        )
        for control in controls
    )


def add_list_validation(ws, cell_range: str, name: str) -> None:
    validation = DataValidation(type="list", formula1=f"={name}", allow_blank=False)
    ws.add_data_validation(validation)
    validation.add(cell_range)


def add_risk_conditional_formats(ws, cell_range: str) -> None:
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=['F4="High (3)"'], fill=PatternFill("solid", fgColor=COLORS["high"])),
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=['F4="Medium (2)"'], fill=PatternFill("solid", fgColor=COLORS["medium"])
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=['F4="Low (1)"'], fill=PatternFill("solid", fgColor=COLORS["low"])),
    )


def add_trigger_conditional_formats(ws, cell_range: str) -> None:
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=['C4="YES"'], fill=PatternFill("solid", fgColor=COLORS["yes_fill"])),
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=['C4="COND"'], fill=PatternFill("solid", fgColor=COLORS["cond_fill"])),
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=['C4="NO"'], fill=PatternFill("solid", fgColor=COLORS["no_fill"])),
    )


def apply_risk_style(cell, risk: Any) -> None:
    risk_text = safe_text(risk)
    fill = {"High": COLORS["high"], "Medium": COLORS["medium"], "Low": COLORS["low"]}.get(
        risk_text
    )
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def apply_trigger_style(cell, trigger: str) -> None:
    fill = {
        "YES": COLORS["yes_fill"],
        "COND": COLORS["cond_fill"],
        "NO": COLORS["no_fill"],
    }.get(trigger)
    color = {
        "YES": COLORS["yes_text"],
        "COND": COLORS["cond_text"],
        "NO": COLORS["no_text"],
    }.get(trigger)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if color:
        cell.font = Font(name="Calibri", size=9, bold=True, color=color)


def title_cell(cell) -> None:
    cell.font = Font(name="Calibri", size=14, bold=True, color=COLORS["white"])
    cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
    cell.alignment = Alignment(horizontal="center")


def style_header_row(ws, row_number: int, column_count: int) -> None:
    for column in range(1, column_count + 1):
        cell = ws.cell(row_number, column)
        cell.font = Font(name="Calibri", size=9, bold=True, color=COLORS["white"])
        cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()


def style_used_range(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if cell.row != 1 and cell.fill.fill_type is None:
                cell.fill = PatternFill("solid", fgColor=COLORS["white"])
            if cell.font == Font():
                cell.font = Font(name="Calibri", size=9, color="000000")
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_number in range(1, ws.max_row + 1):
        ws.row_dimensions[row_number].height = 18 if row_number <= 3 else 48


def thin_border() -> Border:
    side = Side(style="thin", color=COLORS["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def set_widths(ws, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def apply_print_setup(ws, filename: str) -> None:
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.oddFooter.left.text = "&8&F"
    ws.oddFooter.right.text = "&8Page &P of &N"


def join_list(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(safe_text(item) for item in value if safe_text(item))
    return safe_text(value)


def safe_text(value: Any) -> str:
    if value is None or value == "":
        return "[Client To Confirm]"
    return " ".join(str(value).split())


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Render Safe Method document-set JSON to XLSX.")
    parser.add_argument("--document-set-json", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--filename", required=True)
    parser.add_argument("--result-json")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    document_set = json.loads(Path(args.document_set_json).read_text(encoding="utf-8"))
    result = render_xlsx(document_set, Path(args.output), args.filename)
    result_text = json.dumps(result, indent=2)
    if args.result_json:
        Path(args.result_json).write_text(f"{result_text}\n", encoding="utf-8")
    else:
        print(result_text)


if __name__ == "__main__":
    main()
