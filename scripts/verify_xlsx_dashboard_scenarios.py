from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def mutate(workbook: Path) -> None:
    wb = load_workbook(workbook, data_only=False)
    wb["Risk Register"]["G4"] = "High (3)"
    wb["HRCW Register"]["C4"] = "NO"
    wb.save(workbook)


def assert_recalculated(workbook: Path) -> dict[str, object]:
    wb = load_workbook(workbook, data_only=True)
    risk = wb["Risk Register"]
    dashboard = wb["Dashboard"]
    hrcw = wb["HRCW Register"]

    residual_high = sum(1 for row in range(4, risk.max_row + 1) if risk.cell(row, 7).value == "High (3)")
    hrcw_yes = sum(1 for row in range(4, hrcw.max_row + 1) if hrcw.cell(row, 3).value == "YES")
    hrcw_no = sum(1 for row in range(4, hrcw.max_row + 1) if hrcw.cell(row, 3).value == "NO")

    checks = {
        "residual_score_cell": risk["L4"].value,
        "risk_reduction_cell": risk["M4"].value,
        "dashboard_residual_high": dashboard["C13"].value,
        "expected_residual_high": residual_high,
        "dashboard_hrcw_yes": dashboard["C6"].value,
        "expected_hrcw_yes": hrcw_yes,
        "dashboard_hrcw_no": dashboard["C8"].value,
        "expected_hrcw_no": hrcw_no,
    }

    failures = []
    if risk["L4"].value != 3:
        failures.append("Risk Register L4 did not recalculate to residual score 3")
    if risk["M4"].value != 0:
        failures.append("Risk Register M4 did not recalculate risk reduction to 0")
    if dashboard["C13"].value != residual_high:
        failures.append("Dashboard residual High count did not match Risk Register")
    if dashboard["C6"].value != hrcw_yes:
        failures.append("Dashboard HRCW YES count did not match HRCW Register")
    if dashboard["C8"].value != hrcw_no:
        failures.append("Dashboard HRCW NO count did not match HRCW Register")

    return {"status": "pass" if not failures else "fail", "checks": checks, "failures": failures}


def main() -> None:
    parser = argparse.ArgumentParser(description="Phase 5B dashboard scenario helper.")
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--mutate", action="store_true")
    parser.add_argument("--assert-recalculated", action="store_true")
    args = parser.parse_args()

    workbook = Path(args.workbook)
    if args.mutate:
        mutate(workbook)
        return
    if args.assert_recalculated:
        report = assert_recalculated(workbook)
        print(json.dumps(report, indent=2))
        if report["status"] != "pass":
            raise SystemExit(1)
        return
    raise SystemExit("Choose --mutate or --assert-recalculated")


if __name__ == "__main__":
    main()
